import pandas as pd
from collections import defaultdict
import re
import openpyxl
from datetime import datetime as dt
import sys
import numpy as np
import click

@click.command()
@click.argument("school", required=True)
@click.option("--name", "-n", default='Cologne International School', type=str, required=False, help="School Name (please use ' ' to enclose it.)")

def isams_oa_sync(school: str, name:str) -> None:
  print('Starting iSAMS-OA Sync. . .')
  def load_data(school_name):
      oa_file = f'OA ({school_name}).xlsx'
      isams_file = f'iSAMS ({school_name}).xlsx'
      nationality_file = f'CrossReferenceMapping - Nationality - Country.csv'
      grade_year_file = f'grade_year_mapping ({school_name}).csv'
      
      oa_df = pd.read_excel(oa_file)
      isams_df = pd.read_excel(isams_file)
      nationality_country_mapping_df = pd.read_csv(nationality_file)
      grade_year_mapping_dict = pd.read_csv(grade_year_file).groupby('Grade')['Year (NC)'].apply(list).to_dict()
      
      return oa_df, isams_df, nationality_country_mapping_df, grade_year_mapping_dict
      
  school_name = name

  oa_df, isams_df, nationality_country_mapping_df, grade_year_mapping_dict = load_data(school_name)
  mapping_dict = nationality_country_mapping_df.set_index('Title')['ISO'].to_dict()
  grade_year_mapping_dict = {k: v[0] if len(v) == 1 else v for k, v in grade_year_mapping_dict.items()}
  print('Loading all files. . .')
  
  def preprocess_isams(isams, nationality_mapping):
      isams.replace(' ', np.nan, inplace=True)
      isams['Id'] = isams['Forename'] + isams['Surname']    
      def flatten_group(group):
          flattened = {}
          max_i = 0  # Initialize max_i for this group
          for column in ['Primary Contact Email', 'Primary Contact Forename', 'Primary Contact Surname', 'Primary Contact Title', 'Relation Type']:
              unique_values = group[column]
              for i, value in enumerate(unique_values, start=1):
                  flattened[f'{column} {i}'] = value
                  max_i = max(max_i, i)  # Update max_i if this i is larger
          return pd.Series(flattened), max_i
      flattened_results = isams.groupby('Id').apply(flatten_group)
      flattened_df = flattened_results.apply(lambda x: x[0])
      max_i_per_id = flattened_results.apply(lambda x: x[1])
      # Find the overall maximum 'i'
      max_i = max_i_per_id.max()
      merged_df = isams[['Date of Birth', 'Forename', 'Gender', 'Middle Names', 'Preferred Name',
        'Surname', 'School Code', 'Year (NC)', 'Address Type', 'Country',
        'Language', 'Nationality', 'Id']].merge(flattened_df, on='Id', how='inner')
      merged_df_copy = merged_df.drop_duplicates(subset='Id', keep='first').reset_index(drop=True)
      merged_df_copy['Date of Birth'] = pd.to_datetime(merged_df_copy['Date of Birth'], format='%B %d,%Y')
      if not 'Middle Names' in merged_df_copy:
          merged_df_copy['Middle Names'] = np.nan
      def generate_column_names(base_name, max_i):
          """Generates column names up to the maximum index 'i'."""
          return [f"{base_name} {i}" for i in range(1, max_i + 1)]

      forename_columns = generate_column_names('Primary Contact Forename', max_i)
      surname_columns = generate_column_names('Primary Contact Surname', max_i)
      email_columns = generate_column_names('Primary Contact Email', max_i)
      relation_type_columns = generate_column_names('Relation Type', max_i)
      
      merged_df_copy['Parent_first_name_mapped'] = merged_df_copy[forename_columns].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      merged_df_copy['Parent_last_name_mapped'] = merged_df_copy[surname_columns].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      merged_df_copy['Parent_email_mapped'] = merged_df_copy[email_columns].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      merged_df_copy['Parent_relation_mapped'] = merged_df_copy[relation_type_columns].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )

      nationality_columns = merged_df_copy['Nationality'].str.split(r',\s*', expand=True)
      nationality_columns = nationality_columns.rename(columns={i: f'Nationality {i+1}' for i in range(nationality_columns.shape[1])})

      for col in [f'Nationality {i}' for i in range(1, 5)]:
          if col not in nationality_columns.columns:
              nationality_columns[col] = np.nan
          nationality_columns[f'{col}_mapped'] = nationality_columns[col].apply(lambda x: nationality_mapping.get(x) if pd.notna(x) else np.nan)
      merged_df_copy = merged_df_copy.join(nationality_columns)
      merged_df_copy['Nationality_mapped'] = nationality_columns[['Nationality 1_mapped', 'Nationality 2_mapped', 'Nationality 3_mapped', 'Nationality 4_mapped']].apply(
          lambda row: list(set(x for x in row if pd.notna(x))), axis=1
      )
      if not 'Pupil Email Address' in merged_df_copy:
          merged_df_copy['Pupil Email Address'] = np.nan
      return merged_df_copy, max_i

  def preprocess_oa(oa,nationality_mapping, grade_mapping):
      oa['Birth Date'] = pd.to_datetime(oa['Birth Date'], dayfirst=True, format='%d/%m/%y')
      oa['Grade_mapped'] = oa['Grade'].apply(lambda x: [grade_mapping.get(x)] if not isinstance(grade_mapping.get(x), list) else grade_mapping.get(x))
      oa['Parent_first_name_mapped'] = oa[['Parent/Guardian 1 - First Name', 'Parent/Guardian 2 - First Name', 'Parent/Guardian 3 - First Name', 'Parent/Guardian 4 - First Name']].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      oa['Parent_last_name_mapped'] = oa[['Parent/Guardian 1 - Last Name', 'Parent/Guardian 2 - Last Name', 'Parent/Guardian 3 - Last Name', 'Parent/Guardian 4 - Last Name']].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      oa['Parent_email_mapped'] = oa[['Parent/Guardian 1 - Email', 'Parent/Guardian 2 - Email', 'Parent/Guardian 3 - Email', 'Parent/Guardian 4 - Email']].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      oa['Parent_relationship_mapped'] = oa[['Parent/Guardian 1 - Relationship', 'Parent/Guardian 2 - Relationship', 'Parent/Guardian 3 - Relationship', 'Parent/Guardian 4 - Relationship']].apply(
          lambda row: list(set(x.lower() for x in row if pd.notna(x))), axis=1
      )
      oa['Nationality_raw_mapped'] = oa[['Nationality', 'Second Nationality', 'Third Nationality']].apply(
          lambda row: list(x for x in row if pd.notna(x)), axis=1
      )
      nationality_columns = oa['Nationality_raw_mapped'].apply(pd.Series)
      nationality_columns = nationality_columns.rename(columns={0: 'Nationality 1', 1: 'Nationality 2', 2: 'Nationality 3', 3: 'Nationality 4'})
      for col in ['Nationality 1', 'Nationality 2', 'Nationality 3', 'Nationality 4']:
          if col not in nationality_columns.columns:
              nationality_columns[col] = np.nan
          nationality_columns[f'{col}_mapped'] = nationality_columns[col].apply(lambda x: nationality_mapping.get(x) if pd.notna(x) else np.nan)
      oa = oa.join(nationality_columns)
      oa['Nationality_mapped'] = nationality_columns[['Nationality 1_mapped', 'Nationality 2_mapped', 'Nationality 3_mapped', 'Nationality 4_mapped']].apply(
          lambda row: list(set(x for x in row if pd.notna(x))), axis=1
      )
      if not 'Preferred Name' in oa:
          oa['Preferred Names'] = np.nan
      return oa

  isams_df_copy, max_i = preprocess_isams(isams_df.copy(), mapping_dict)
  oa_df_copy = preprocess_oa(oa_df.copy(),mapping_dict, grade_year_mapping_dict)
  isams_df_copy = isams_df_copy.add_prefix('isams_')
  oa_df_copy = oa_df_copy.add_prefix('oa_')
  oa_df_copy_enrolled = oa_df_copy[oa_df_copy['oa_Student Status'] == 'Enrolled'].reset_index(drop=True)
  oa_df_copy_others = oa_df_copy.drop(oa_df_copy_enrolled.index).reset_index(drop=True)
  print('Preprocessing files. . .')
  print('Starting merge sequence. . .')
  
  #first merge (isams with oa_enrolled)
  merged_df_1st = isams_df_copy[isams_df_copy['isams_School Code'].notna()].merge(oa_df_copy_enrolled[oa_df_copy_enrolled['oa_Student ID'].notna()],left_on='isams_School Code', right_on='oa_Student ID', how='inner')
  matched_id_isams = merged_df_1st['isams_School Code'].unique()
  matched_id_oa = merged_df_1st['oa_Student ID'].unique()
  isams_for_2nd_round = isams_df_copy[(isams_df_copy['isams_School Code'].isna()) | ~(isams_df_copy['isams_School Code'].isin(matched_id_isams))].reset_index(drop=True)
  leftover_oa = oa_df_copy_enrolled[(oa_df_copy_enrolled['oa_Student ID'].isna() | ~(oa_df_copy_enrolled['oa_Student ID'].isin(matched_id_oa)))].reset_index(drop=True)

  merged_df_2nd = isams_for_2nd_round[isams_for_2nd_round['isams_School Code'].notna()].merge(oa_df_copy_others[oa_df_copy_others['oa_Student ID'].notna()],left_on='isams_School Code', right_on='oa_Student ID', how='inner')
  matched_id_isams_2nd = merged_df_2nd['isams_School Code'].unique()
  leftover_isams = isams_for_2nd_round[(isams_for_2nd_round['isams_School Code'].isna()) | ~(isams_for_2nd_round['isams_School Code'].isin(matched_id_isams_2nd))].reset_index(drop=True)

  merged_df_1st['Note'] = ''
  merged_df_2nd['Note'] = 'Enrolled in iSAMS; Not Enrolled in OA'
  leftover_oa['Note'] = 'Student not in iSAMS'
  leftover_isams['Note'] = 'Student not in OA'
  merged_df = pd.concat([merged_df_1st, merged_df_2nd])
  print('Merge sequence completed.')
  
  def add_comparison_columns(df):
      comparison_mappings = {
          'is_same_id': ('isams_School Code','oa_Student ID'),
          'is_same_first_name': ('isams_Forename','oa_First Name'),
          'is_same_last_name': ('isams_Surname','oa_Last Name'),
          'is_same_middle_name': ('isams_Middle Names','oa_Middle Name(s)'),
          'is_same_email': ('isams_Pupil Email Address', 'oa_Email'),
          'is_same_preferred_name': ('isams_Preferred Name','oa_Preferred Names'),
          'is_same_grade_year': ('isams_Year (NC)','oa_Grade_mapped'),
          'is_same_date_of_birth': ('isams_Date of Birth','oa_Birth Date')
      }

      for new_col, (col1, col2) in comparison_mappings.items():
          if col1 in df.columns and col2 in df.columns:
              if new_col == 'is_same_grade_year':
                  # Special handling for grade year comparison
                  df[new_col] = df.apply(lambda row: row[col1] in row[col2], axis=1)
              elif new_col == 'is_same_id':
                  df[new_col] = df[col1] == df[col2]
              else:
                  # Standard comparison for other columns
                  df[new_col] = np.where(
                      df[col1].isna() & df[col2].isna(),
                      True,  # Set to True if both are NaN
                      df[col1].astype(str).str.lower() == df[col2].astype(str).str.lower()
                  )
      # Checking nationalities
      oa_nationality_cols = [f'oa_Nationality {i}_mapped' for i in range(1, 5)]
      isams_nationality_cols = [f'isams_Nationality {i}_mapped' for i in range(1, 5)]

      # Check each OA nationality against all iSAMS nationalities
      for i, oa_nat_col in enumerate(oa_nationality_cols, start=1):
          df[f'is_same_nationality_{i}_from_oa'] = df.apply(
              lambda row: row[oa_nat_col] in row['isams_Nationality_mapped'] if pd.notna(row[oa_nat_col]) else False, 
              axis=1)

      # Check each iSAMS nationality against all OA nationalities
      for i, isams_nat_col in enumerate(isams_nationality_cols, start=1):
          df[f'is_same_nationality_{i}_from_isams'] = df.apply(
              lambda row: row[isams_nat_col] in row['oa_Nationality_mapped'] if pd.notna(row[isams_nat_col]) else False, 
              axis=1)
      return df

  def add_parents_comparison_columns(df,max_i):
      # Checking Parent First Name
      oa_parent_first_name_cols = [f'oa_Parent/Guardian {i} - First Name' for i in range(1, 5)]
      isams_parent_first_name_cols = [f'isams_Primary Contact Forename {i}' for i in range(1, max_i + 1)]

      for i, oa_pfn_col in enumerate(oa_parent_first_name_cols, start=1):
          df[f'is_same_parent_first_name_{i}_from_oa'] = df.apply(
              lambda row: row[oa_pfn_col].lower() in row['isams_Parent_first_name_mapped']
                  if pd.notna(row[oa_pfn_col])
                      else False, 
              axis=1)

      for i, isams_pfn_col in enumerate(isams_parent_first_name_cols, start=1):
          df[f'is_same_parent_first_name_{i}_from_isams'] = df.apply(
              lambda row: row[isams_pfn_col].lower() in row['oa_Parent_first_name_mapped'] 
                  if pd.notna(row[isams_pfn_col])
                  else False, 
              axis=1)
      
      oa_parent_last_name_cols = [f'oa_Parent/Guardian {i} - Last Name' for i in range(1, 5)]
      isams_parent_last_name_cols = [f'isams_Primary Contact Surname {i}' for i in range(1, max_i + 1)]

      for i, oa_pln_col in enumerate(oa_parent_last_name_cols, start=1):
          df[f'is_same_parent_last_name_{i}_from_oa'] = df.apply(
              lambda row: row[oa_pln_col].lower() in row['isams_Parent_last_name_mapped']
                  if pd.notna(row[oa_pln_col])
                      else False, 
              axis=1)

      for i, isams_pln_col in enumerate(isams_parent_last_name_cols, start=1):
          df[f'is_same_parent_last_name_{i}_from_isams'] = df.apply(
              lambda row: row[isams_pln_col].lower() in row['oa_Parent_last_name_mapped'] 
                  if pd.notna(row[isams_pln_col])
                  else False, 
              axis=1)
          
      oa_parent_email_cols = [f'oa_Parent/Guardian {i} - Email' for i in range(1, 5)]
      isams_parent_email_cols = [f'isams_Primary Contact Email {i}' for i in range(1, max_i + 1)]

      for i, oa_pe_col in enumerate(oa_parent_email_cols, start=1):
          df[f'is_same_parent_email_{i}_from_oa'] = df.apply(
              lambda row: row[oa_pe_col].lower() in row['isams_Parent_email_mapped']
                  if pd.notna(row[oa_pe_col])
                      else False, 
              axis=1)

      for i, isams_pe_col in enumerate(isams_parent_email_cols, start=1):
          df[f'is_same_parent_email_{i}_from_isams'] = df.apply(
              lambda row: row[isams_pe_col].lower() in row['oa_Parent_email_mapped'] 
                  if pd.notna(row[isams_pe_col])
                  else False, 
              axis=1)
          
      oa_parent_relationship_cols = [f'oa_Parent/Guardian {i} - Relationship' for i in range(1, 5)]
      isams_parent_relationship_cols = [f'isams_Relation Type {i}' for i in range(1, max_i + 1)]

      for i, oa_pr_col in enumerate(oa_parent_relationship_cols, start=1):
          df[f'is_same_parent_relationship_{i}_from_oa'] = df.apply(
              lambda row: row[oa_pr_col].lower() in row['isams_Parent_relation_mapped']
                  if pd.notna(row[oa_pr_col])
                      else False, 
              axis=1)

      for i, isams_pr_col in enumerate(isams_parent_relationship_cols, start=1):
          df[f'is_same_parent_relationship_{i}_from_isams'] = df.apply(
              lambda row: row[isams_pr_col].lower() in row['oa_Parent_relationship_mapped'] 
                  if pd.notna(row[isams_pr_col])
                  else False, 
              axis=1)
      return df
  print('Comparing. . .')
  
  #Analyse merged
  merged_df_copy = add_comparison_columns(merged_df.copy())
  merged_df_copy = add_parents_comparison_columns(merged_df_copy, max_i)
  export_df = pd.concat([merged_df_copy,leftover_isams,leftover_oa], axis=0, ignore_index=True)

  new_order = [
      'Note',
      'isams_School Code', 'isams_Pupil Email Address', 'isams_Date of Birth',
      'isams_Forename', 'isams_Middle Names', 'isams_Surname', 'isams_Preferred Name',
      'isams_Year (NC)', 'isams_Gender',
      'isams_Nationality 1', 'isams_Nationality 2',
      'isams_Nationality 3', 'isams_Nationality 4',
  ]

  # Dynamically add 'Primary Contact' entries based on max_i
  for i in range(1, max_i + 1):
      new_order.extend([
          f'isams_Primary Contact Forename {i}', f'isams_Primary Contact Surname {i}',
          f'isams_Primary Contact Email {i}', f'isams_Relation Type {i}',
      ])
  new_order.extend([
      'oa_Student Status', 'oa_Student ID', 'oa_Email', 'oa_Birth Date',
      'oa_First Name', 'oa_Middle Name(s)', 'oa_Last Name', 'oa_Preferred Names',
      'oa_Grade', 'oa_Gender',

      'oa_Nationality 1', 'oa_Nationality 2',
      'oa_Nationality 3', 'oa_Nationality 4',

      'oa_Parent/Guardian 1 - First Name', 'oa_Parent/Guardian 1 - Last Name',
      'oa_Parent/Guardian 1 - Email', 'oa_Parent/Guardian 1 - Relationship',

      'oa_Parent/Guardian 2 - First Name', 'oa_Parent/Guardian 2 - Last Name',
      'oa_Parent/Guardian 2 - Email', 'oa_Parent/Guardian 2 - Relationship',

      'oa_Parent/Guardian 3 - First Name', 'oa_Parent/Guardian 3 - Last Name',
      'oa_Parent/Guardian 3 - Email', 'oa_Parent/Guardian 3 - Relationship',

      'oa_Parent/Guardian 4 - First Name', 'oa_Parent/Guardian 4 - Last Name',
      'oa_Parent/Guardian 4 - Email', 'oa_Parent/Guardian 4 - Relationship',

      'isams_Address Type', 'isams_Country', 'isams_Language',
      'isams_Nationality', 'isams_Id', 

      'isams_Parent_first_name_mapped', 'isams_Parent_last_name_mapped',
      'isams_Parent_email_mapped', 'isams_Parent_relation_mapped',
      'isams_Nationality 1_mapped', 'isams_Nationality 2_mapped',
      'isams_Nationality 3_mapped', 'isams_Nationality 4_mapped',
      'isams_Nationality_mapped',

      'oa_OpenApply ID', 'oa_OpenApply URL',
      'oa_Nationality', 'oa_Second Nationality', 'oa_Third Nationality', 
      'oa_Parent/Guardian 1 - Parent OpenApply ID', 'oa_Parent/Guardian 2 - Parent OpenApply ID',
      'oa_Parent/Guardian 3 - Parent OpenApply ID', 'oa_Parent/Guardian 4 - Parent OpenApply ID',
      'oa_Grade_mapped', 'oa_Parent_first_name_mapped', 'oa_Parent_last_name_mapped',
      'oa_Parent_email_mapped', 'oa_Parent_relationship_mapped', 'oa_Nationality 1_mapped',
      'oa_Nationality 2_mapped', 'oa_Nationality 3_mapped',
      'oa_Nationality 4_mapped', 'oa_Nationality_mapped', 
      'is_same_id', 'is_same_first_name', 'is_same_last_name', 'is_same_preferred_name',
      'is_same_middle_name', 'is_same_grade_year', 'is_same_date_of_birth', 'is_same_email',
      'is_same_nationality_1_from_oa', 'is_same_nationality_2_from_oa',
      'is_same_nationality_3_from_oa', 'is_same_nationality_4_from_oa',
      'is_same_nationality_1_from_isams', 'is_same_nationality_2_from_isams',
      'is_same_nationality_3_from_isams', 'is_same_nationality_4_from_isams',
      'is_same_parent_first_name_1_from_oa', 'is_same_parent_first_name_2_from_oa',
      'is_same_parent_first_name_3_from_oa', 'is_same_parent_first_name_4_from_oa'
  ])

  for i in range(1, max_i + 1):
      new_order.extend([
          f'is_same_parent_first_name_{i}_from_isams',
      ])

  new_order.extend([
      'is_same_parent_last_name_1_from_oa', 'is_same_parent_last_name_2_from_oa',
      'is_same_parent_last_name_3_from_oa', 'is_same_parent_last_name_4_from_oa',
  ])

  for i in range(1, max_i + 1):
      new_order.extend([
          f'is_same_parent_last_name_{i}_from_isams',
      ])

  new_order.extend([
      'is_same_parent_email_1_from_oa', 'is_same_parent_email_2_from_oa',
      'is_same_parent_email_3_from_oa', 'is_same_parent_email_4_from_oa',
  ])

  for i in range(1, max_i + 1):
      new_order.extend([
          f'is_same_parent_email_{i}_from_isams',
      ])

  new_order.extend([
      'is_same_parent_relationship_1_from_oa', 'is_same_parent_relationship_2_from_oa',
      'is_same_parent_relationship_3_from_oa', 'is_same_parent_relationship_4_from_oa',
  ])

  for i in range(1, max_i + 1):
      new_order.extend([
          f'is_same_parent_relationship_{i}_from_isams',
      ])

  def update_note(row):
      if 'Only' in row['Note']:
          return row
      
      if pd.isna(row['Note']) or row['Note'] == '':
          notes = []
      else:
          notes = [row['Note']]

      comparison_mappings = {
          'is_same_id': ('iSAMS School Code', 'OA Student ID'),
          'is_same_email': ('iSAMS Pupil Email Address', 'OA Email'),
          'is_same_first_name': ('iSAMS Forename', 'OA First Name'),
          'is_same_last_name': ('iSAMS Surname', 'OA Last Name'),
          'is_same_middle_name': ('iSAMS Middle Names', 'OA Middle Name(s)'),
          'is_same_preferred_name': ('iSAMS Preferred Name', 'OA Preferred Names'),
          'is_same_grade_year': ('iSAMS Year (NC)', 'OA Grade_mapped'),
          'is_same_date_of_birth': ('iSAMS Date of Birth', 'OA Birth Date')
      }
      
      for comparison, (col1, col2) in comparison_mappings.items():
          if comparison in row and not row[comparison]:
              # Check if either of the columns is empty
              if pd.isna(row[col1]) or pd.isna(row[col2]):
                  missing_column = col1 if pd.isna(row[col1]) else col2
                  friendly_missing = missing_column.split()[0]
                  friendly_name = comparison.replace('is_same_', '').replace('_', ' ').title()
                  notes.append(f'Missing {friendly_name} in {friendly_missing}')
              else:
                  friendly_name = comparison.replace('is_same_', '').replace('_', ' ').title()
                  notes.append(f'Conflict {friendly_name}')
      
      for i in range(1, 5):
          isams_nationality = pd.isna(row.get(f'iSAMS Nationality {i}'))
          oa_nationality = pd.isna(row.get(f'OA Nationality {i}'))
          is_same_nationality_isams = row.get(f'is_same_nationality_{i}_from_isams') is False
          is_same_nationality_oa = row.get(f'is_same_nationality_{i}_from_oa') is False

          if ((is_same_nationality_oa) and (not oa_nationality)) or ((is_same_nationality_isams) and (not isams_nationality)):
              notes.append(f'Conflict Nationality')
              break  # Only add the note once
      
      attribute_mappings = [
          ("Email", "Primary Contact Email", "Conflict Parent Email"),
          ("First Name", "Primary Contact Forename", "Conflict Parent First Name"),
          ("Last Name", "Primary Contact Surname", "Conflict Parent Last Name"),
          ("Relationship", "Relation Type", "Conflict Parent Relationship"),
      ]

      # Iterate over the attribute mappings
      for i in range(1, 5):
          for attribute, source_prefix, note_message in attribute_mappings:
              if attribute == "Email":
                  isams_parent_emails = row.get('iSAMS Parent_email_mapped')
                  oa_email = row.get(f'OA Parent/Guardian {i} - {attribute}')
                  if not isinstance(isams_parent_emails, list):
                      isams_parent_emails = []
                  if isinstance(oa_email, str):
                      if not oa_email in isams_parent_emails:
                          notes.append(f'Parent {i} from OA is missing in iSAMS')
                          break
              oa_parent_attribute = pd.isna(row.get(f'OA Parent/Guardian {i} - {attribute}'))
              is_same_parent_attribute_oa = row.get(f"is_same_{' '.join(note_message.split()[1:]).lower().replace(' ','_')}_{i}_from_oa") is False
              if ((is_same_parent_attribute_oa) and (not oa_parent_attribute)):
                  if not note_message in notes:
                      notes.append(note_message)
                  
      for i in range(1, max_i + 1):    
          # Check for conflicts in iSAMS range
          for attribute, source_prefix, note_message in attribute_mappings:
              if attribute == "Email":
                  # Initialize OA email list or set to empty if NaN
                  oa_parent_emails = row.get('OA Parent_email_mapped')
                  isams_email = row.get(f'iSAMS {source_prefix} {i}')
                  if not isinstance(oa_parent_emails, list):
                      oa_parent_emails = []
                  if isinstance(isams_email, str):
                      if not isams_email in oa_parent_emails:
                          notes.append(f'Parent {i} from iSAMS is missing in OA')
                          break
              isams_parent_attribute = pd.isna(row.get(f'iSAMS {source_prefix} {i}'))
              is_same_parent_attribute_isams = row.get(f"is_same_{' '.join(note_message.split()[1:]).lower().replace(' ','_')}_{i}_from_isams") is False

              if is_same_parent_attribute_isams and not isams_parent_attribute:
                  if not note_message in notes:
                      notes.append(note_message)
      if notes:
          row['Note'] = ', '.join(notes)
      return row
  print('Comparison process done.')
  
  export_df_copy = export_df[new_order].copy()
  export_df_copy = export_df_copy.rename(columns=lambda x: x.replace('oa_', 'OA ').replace('isams_', 'iSAMS '))
  export_df_copy = export_df_copy.apply(update_note, axis=1)
  export_df_copy['OA Birth Date'] = pd.to_datetime(export_df_copy['OA Birth Date']).dt.strftime('%d %B, %Y')
  export_df_copy['iSAMS Date of Birth'] = pd.to_datetime(export_df_copy['iSAMS Date of Birth']).dt.strftime('%d %B, %Y')
  export_df_copy = export_df_copy.sort_index()

  columns_to_hide = [
          'iSAMS Address Type', 'iSAMS Country', 'iSAMS Language',
          'iSAMS Nationality', 'iSAMS Id', 
      
          'iSAMS Parent_first_name_mapped', 'iSAMS Parent_last_name_mapped',
          'iSAMS Parent_email_mapped', 'iSAMS Parent_relation_mapped',
          'iSAMS Nationality 1_mapped', 'iSAMS Nationality 2_mapped',
          'iSAMS Nationality 3_mapped', 'iSAMS Nationality 4_mapped',
          'iSAMS Nationality_mapped',
      
          'OA OpenApply ID', 'OA OpenApply URL', 
          'OA Nationality', 'OA Second Nationality', 'OA Third Nationality', 
          'OA Parent/Guardian 1 - Parent OpenApply ID', 'OA Parent/Guardian 2 - Parent OpenApply ID',
          'OA Parent/Guardian 3 - Parent OpenApply ID', 'OA Parent/Guardian 4 - Parent OpenApply ID',
          'OA Grade_mapped', 'OA Parent_first_name_mapped', 'OA Parent_last_name_mapped',
          'OA Parent_email_mapped', 'OA Parent_relationship_mapped', 'OA Nationality 1_mapped',
          'OA Nationality 2_mapped', 'OA Nationality 3_mapped',
          'OA Nationality 4_mapped', 'OA Nationality_mapped', 
          'is_same_id', 'is_same_first_name', 'is_same_last_name', 'is_same_preferred_name',
          'is_same_middle_name', 'is_same_grade_year', 'is_same_date_of_birth', 'is_same_email',
          'is_same_nationality_1_from_oa', 'is_same_nationality_2_from_oa',
          'is_same_nationality_3_from_oa', 'is_same_nationality_4_from_oa',
          'is_same_nationality_1_from_isams', 'is_same_nationality_2_from_isams',
          'is_same_nationality_3_from_isams', 'is_same_nationality_4_from_isams',
          'is_same_parent_first_name_1_from_oa', 'is_same_parent_first_name_2_from_oa',
          'is_same_parent_first_name_3_from_oa', 'is_same_parent_first_name_4_from_oa',
          'is_same_parent_last_name_1_from_oa', 'is_same_parent_last_name_2_from_oa',
          'is_same_parent_last_name_3_from_oa', 'is_same_parent_last_name_4_from_oa',
          'is_same_parent_email_1_from_oa', 'is_same_parent_email_2_from_oa',
          'is_same_parent_email_3_from_oa', 'is_same_parent_email_4_from_oa',
          'is_same_parent_relationship_1_from_oa', 'is_same_parent_relationship_2_from_oa',
          'is_same_parent_relationship_3_from_oa', 'is_same_parent_relationship_4_from_oa',
  ]

  for i in range(1, max_i + 1):
      columns_to_hide.extend([
          f'is_same_parent_first_name_{i}_from_isams', f'is_same_parent_last_name_{i}_from_isams',
          f'is_same_parent_email_{i}_from_isams', f'is_same_parent_relationship_{i}_from_isams',
      ])

  def highlight_columns(row):
      if 'not in' in row['Note']:
          return {col: '' for col in row.index}  # Return default style for all columns
      # Define the comparison mappings
      comparison_mappings = {
          'is_same_id': ('iSAMS School Code', 'OA Student ID'),
          'is_same_email': ('iSAMS Pupil Email Address', 'OA Email'),
          'is_same_first_name': ('iSAMS Forename', 'OA First Name'),
          'is_same_last_name': ('iSAMS Surname', 'OA Last Name'),
          'is_same_middle_name': ('iSAMS Middle Names', 'OA Middle Name(s)'),
          'is_same_preferred_name': ('iSAMS Preferred Name', 'OA Preferred Names'),
          'is_same_grade_year': ('iSAMS Year (NC)', 'OA Grade_mapped'),
          'is_same_date_of_birth': ('iSAMS Date of Birth', 'OA Birth Date')
      }

      styles = {}
      for comparison, (col1, col2) in comparison_mappings.items():
          if comparison in row and not row[comparison]:
              styles[col1] = 'color: #f40404; font-weight: bold'
              styles[col2] = 'color: #f40404; font-weight: bold'
      for i in range(1, 5):
          if not row.get(f'is_same_nationality_{i}_from_oa', True):
              styles[f'OA Nationality {i}'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_nationality_{i}_from_isams', True):
              styles[f'iSAMS Nationality {i}'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_first_name_{i}_from_oa', True) and (row[f'is_same_parent_first_name_{i}_from_oa'] != ''):
              styles[f'OA Parent/Guardian {i} - First Name'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_last_name_{i}_from_oa', True) and (row[f'is_same_parent_last_name_{i}_from_oa'] != ''):
              styles[f'OA Parent/Guardian {i} - Last Name'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_relationship_{i}_from_oa', True) and (row[f'is_same_parent_relationship_{i}_from_oa'] != ''):
              styles[f'OA Parent/Guardian {i} - Relationship'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_email_{i}_from_oa', True) and (row[f'is_same_parent_email_{i}_from_oa'] != ''):
              styles[f'OA Parent/Guardian {i} - Email'] = 'color: #f40404; font-weight: bold'
      for i in range(1,max_i+1):
          if not row.get(f'is_same_parent_first_name_{i}_from_isams', True) and (row[f'is_same_parent_first_name_{i}_from_isams'] != ''):
              styles[f'iSAMS Primary Contact Forename {i}'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_last_name_{i}_from_isams', True) and (row[f'is_same_parent_last_name_{i}_from_isams'] != ''):
              styles[f'iSAMS Primary Contact Surname {i}'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_relationship_{i}_from_isams', True) and (row[f'is_same_parent_relationship_{i}_from_isams'] != ''):
              styles[f'iSAMS Relation Type {i}'] = 'color: #f40404; font-weight: bold'
          if not row.get(f'is_same_parent_email_{i}_from_isams', True) and (row[f'is_same_parent_email_{i}_from_isams'] != ''):
              styles[f'iSAMS Primary Contact Email {i}'] = 'color: #f40404; font-weight: bold'
      for col in row.index:
          if col not in styles:
              styles[col] = ''
      return pd.Series(styles)

  def highlight_rows(row):
      styles = [''] * len(row)  # Start with default style
      oa_student_id_idx = row.index.get_loc('OA Student ID') - 1  # Get the index of 'MB Student ID' column
      if 'Student not in OA' in row['Note']:
          styles[oa_student_id_idx:] = ['background-color: #07AAE2'] * (len(row) - oa_student_id_idx)
      elif 'Student not in iSAMS' in row['Note']:
          styles[:oa_student_id_idx] = ['background-color: #FFC000'] * oa_student_id_idx

      return styles

  export_df_style = (export_df_copy.fillna('').style.apply(highlight_columns, axis=1).apply(highlight_rows, axis=1))
  export_df_id_conflict = export_df_copy[export_df_copy['is_same_id'] == False].copy().reset_index(drop=True)[['Note', 'iSAMS School Code', 'OA Student ID', 'is_same_id']]
  export_df_id_conflict['Note'] = export_df_id_conflict['Note'].apply(
      lambda x: ', '.join(note for note in x.split(', ') if 'ID' in note)
  )
  export_df_id_conflict_style = export_df_id_conflict.fillna('').style.apply(highlight_columns, axis=1)

  export_df_email_conflict = export_df_copy[export_df_copy['is_same_email'] == False].copy().reset_index(drop=True)[['Note', 'iSAMS Pupil Email Address', 'OA Email', 'is_same_email']]
  export_df_email_conflict['Note'] = export_df_email_conflict['Note'].apply(
      lambda x: ', '.join(note for note in x.split(', ') if 'Email' in note)
  )
  export_df_email_conflict_style = export_df_email_conflict.fillna('').style.apply(highlight_columns, axis=1)

  export_df_dob_conflict = export_df_copy[
      export_df_copy['is_same_date_of_birth'] == False
  ].copy().reset_index(drop=True)[[
      'Note', 
      'iSAMS School Code', 'iSAMS Pupil Email Address', 'iSAMS Date of Birth',
      'OA Student ID', 'OA Email', 'OA Birth Date', 
      'is_same_date_of_birth']]
  export_df_dob_conflict['Note'] = export_df_dob_conflict['Note'].apply(
      lambda x: ', '.join(note for note in x.split(', ') if 'Birth' in note)
  )
  export_df_dob_conflict_style = export_df_dob_conflict.fillna('').style.apply(highlight_columns, axis=1)

  export_df_name_conflict = export_df_copy[
      (export_df_copy['is_same_first_name'] == False) | 
      (export_df_copy['is_same_middle_name'] == False) | 
      (export_df_copy['is_same_last_name'] == False) | 
      (export_df_copy['is_same_preferred_name'] == False)
  ].copy().reset_index(drop=True)[[
      'Note',
      'iSAMS School Code', 'iSAMS Pupil Email Address',
      'iSAMS Forename', 'iSAMS Middle Names', 'iSAMS Surname', 'iSAMS Preferred Name',
      'OA Student ID', 'OA Email',
      'OA First Name', 'OA Middle Name(s)', 'OA Last Name', 'OA Preferred Names', 
      'is_same_first_name', 'is_same_middle_name', 'is_same_last_name', 'is_same_preferred_name'
  ]]
  export_df_name_conflict['Note'] = export_df_name_conflict['Note'].apply(
      lambda x: ', '.join(note for note in x.split(', ') if 'Name' in note)
  )
  export_df_name_conflict_style = export_df_name_conflict.fillna('').style.apply(highlight_columns, axis=1)

  export_df_nationality_conflict = export_df_copy[
      ((export_df_copy['is_same_nationality_1_from_isams'] == False) & export_df_copy['iSAMS Nationality 1'].notna()) |
      ((export_df_copy['is_same_nationality_2_from_isams'] == False) & export_df_copy['iSAMS Nationality 2'].notna()) |
      ((export_df_copy['is_same_nationality_3_from_isams'] == False) & export_df_copy['iSAMS Nationality 3'].notna()) |
      ((export_df_copy['is_same_nationality_4_from_isams'] == False) & export_df_copy['iSAMS Nationality 4'].notna()) |
      ((export_df_copy['is_same_nationality_1_from_oa'] == False) & export_df_copy['OA Nationality 1'].notna()) |
      ((export_df_copy['is_same_nationality_2_from_oa'] == False) & export_df_copy['OA Nationality 2'].notna()) |
      ((export_df_copy['is_same_nationality_3_from_oa'] == False) & export_df_copy['OA Nationality 3'].notna()) |
      ((export_df_copy['is_same_nationality_4_from_oa'] == False) & export_df_copy['OA Nationality 4'].notna())
  ].copy().reset_index(drop=True)[
      ['Note',
      'iSAMS School Code', 'iSAMS Pupil Email Address',
      'iSAMS Nationality 1', 'iSAMS Nationality 2', 'iSAMS Nationality 3', 'iSAMS Nationality 4',
      'OA Student ID', 'OA Email',
      'OA Nationality 1', 'OA Nationality 2', 'OA Nationality 3', 'OA Nationality 4',
      'is_same_nationality_1_from_isams', 'is_same_nationality_2_from_isams',
      'is_same_nationality_3_from_isams', 'is_same_nationality_4_from_isams',
      'is_same_nationality_1_from_oa', 'is_same_nationality_2_from_oa', 
      'is_same_nationality_3_from_oa', 'is_same_nationality_4_from_oa', 
      ]]
  export_df_nationality_conflict['Note'] = export_df_nationality_conflict['Note'].apply(
      lambda x: ', '.join(note for note in x.split(', ') if 'Nationality' in note)
  )
  export_df_nationality_conflict_style = export_df_nationality_conflict.fillna('').style.apply(highlight_columns, axis=1)

  conditions = []

  # Generate conditions for each 'is_same_parent' column dynamically based on max_i
  for i in range(1, max_i + 1):
      conditions.append((export_df_copy[f'is_same_parent_first_name_{i}_from_isams'] == False) & export_df_copy[f'iSAMS Primary Contact Forename {i}'].notna())
      conditions.append((export_df_copy[f'is_same_parent_last_name_{i}_from_isams'] == False) & export_df_copy[f'iSAMS Primary Contact Surname {i}'].notna())
      conditions.append((export_df_copy[f'is_same_parent_relationship_{i}_from_isams'] == False) & export_df_copy[f'iSAMS Relation Type {i}'].notna()) 
      conditions.append((export_df_copy[f'is_same_parent_email_{i}_from_isams'] == False) & export_df_copy[f'iSAMS Primary Contact Email {i}'].notna())

  for i in range(1, 5):
      conditions.append((export_df_copy[f'is_same_parent_first_name_{i}_from_oa'] == False) & export_df_copy[f'OA Parent/Guardian {i} - First Name'].notna())
      conditions.append((export_df_copy[f'is_same_parent_last_name_{i}_from_oa'] == False) & export_df_copy[f'OA Parent/Guardian {i} - Last Name'].notna())
      conditions.append((export_df_copy[f'is_same_parent_relationship_{i}_from_oa'] == False) & export_df_copy[f'OA Parent/Guardian {i} - Relationship'].notna()) 
      conditions.append((export_df_copy[f'is_same_parent_email_{i}_from_oa'] == False) & export_df_copy[f'OA Parent/Guardian {i} - Email'].notna())
      
  # Combine all conditions using 'or'
  combined_condition = conditions[0]
  for condition in conditions[1:]:
      combined_condition |= condition

  columns_parent_conflict = ['Note', 'iSAMS School Code', 'iSAMS Pupil Email Address',]
  for i in range(1, max_i + 1):
      columns_parent_conflict.extend([
          f'iSAMS Primary Contact Forename {i}', f'iSAMS Primary Contact Surname {i}',
          f'iSAMS Primary Contact Email {i}', f'iSAMS Relation Type {i}',
      ])
  columns_parent_conflict.extend(['OA Student ID', 'OA Email',])
  for i in range(1, 5):
      columns_parent_conflict.extend([
          f'OA Parent/Guardian {i} - First Name', f'OA Parent/Guardian {i} - Last Name',
          f'OA Parent/Guardian {i} - Email', f'OA Parent/Guardian {i} - Relationship',
      ])
  for i in range(1, max_i + 1):
      columns_parent_conflict.extend([
          f'is_same_parent_first_name_{i}_from_isams',
          f'is_same_parent_last_name_{i}_from_isams',
          f'is_same_parent_email_{i}_from_isams',
          f'is_same_parent_relationship_{i}_from_isams',
      ])
  for i in range(1, 5):
      columns_parent_conflict.extend([
          f'is_same_parent_first_name_{i}_from_oa',
          f'is_same_parent_last_name_{i}_from_oa',
          f'is_same_parent_email_{i}_from_oa',
          f'is_same_parent_relationship_{i}_from_oa',
      ])
  export_df_parent_conflict = export_df_copy[combined_condition].copy().reset_index(drop=True)[
      columns_parent_conflict
  ]
  export_df_parent_conflict['Note'] = export_df_parent_conflict['Note'].apply(
      lambda x: ', '.join(note for note in x.split(', ') if 'Parent' in note)
  )
  export_df_parent_conflict_style = export_df_parent_conflict.fillna('').style.apply(highlight_columns, axis=1)

  oa_student_id_idx = export_df_copy.loc[0].index.get_loc('OA Student ID')  # Get the index of 'MB Student ID' column
  print('Exporting. . .')
  
  with pd.ExcelWriter(f"isams_oa_analysis_{school_name}_{dt.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx", engine='openpyxl') as writer:
      export_df_style.to_excel(writer, index=False, sheet_name='All Comparison')
      # Get the openpyxl objects
      worksheet = writer.sheets['All Comparison']
      # Coloring headers
      blue_fill = openpyxl.styles.PatternFill(start_color='07AAE2', end_color='07AAE2', fill_type='solid')
      orange_fill = openpyxl.styles.PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
      
      for col in range(2, worksheet.max_column + 1):
          cell = worksheet.cell(row=1, column=col)
          if (col > 1) & (col < oa_student_id_idx):  # adjust according to where 'Student ID' is now
              cell.fill = orange_fill
          elif (col >= oa_student_id_idx):
              cell.fill = blue_fill

      for row in worksheet.iter_rows(min_row=2, max_col=1, max_row=worksheet.max_row):
          for cell in row:
              if cell.value == '':
                  worksheet.row_dimensions[cell.row].hidden = True

      for column in columns_to_hide:
          # print(final_df_copy.columns)
          col_idx = export_df_copy.columns.get_loc(column) + 1  # +1 because Excel is 1-indexed
          worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].hidden = True
          
      export_df_id_conflict_style.to_excel(writer, index=False, sheet_name='ID Conflict')
      worksheet = writer.sheets['ID Conflict']
      worksheet.cell(row=1, column=2).fill = orange_fill
      worksheet.cell(row=1, column=3).fill = blue_fill
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(4)].hidden = True
      
      export_df_email_conflict_style.to_excel(writer, index=False, sheet_name='Email Conflict')
      worksheet = writer.sheets['Email Conflict']
      worksheet.cell(row=1, column=2).fill = orange_fill
      worksheet.cell(row=1, column=3).fill = blue_fill
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(4)].hidden = True
      
      export_df_dob_conflict_style.to_excel(writer, index=False, sheet_name='DOB Conflict')
      worksheet = writer.sheets['DOB Conflict']
      worksheet.cell(row=1, column=2).fill = orange_fill
      worksheet.cell(row=1, column=3).fill = orange_fill
      worksheet.cell(row=1, column=4).fill = orange_fill
      worksheet.cell(row=1, column=5).fill = blue_fill
      worksheet.cell(row=1, column=6).fill = blue_fill
      worksheet.cell(row=1, column=7).fill = blue_fill
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(8)].hidden = True
      
      export_df_name_conflict_style.to_excel(writer, index=False, sheet_name='Name Conflict')
      worksheet = writer.sheets['Name Conflict']
      worksheet.cell(row=1, column=2).fill = orange_fill
      worksheet.cell(row=1, column=3).fill = orange_fill
      worksheet.cell(row=1, column=4).fill = orange_fill
      worksheet.cell(row=1, column=5).fill = orange_fill
      worksheet.cell(row=1, column=6).fill = orange_fill
      worksheet.cell(row=1, column=7).fill = orange_fill
      worksheet.cell(row=1, column=8).fill = blue_fill
      worksheet.cell(row=1, column=9).fill = blue_fill
      worksheet.cell(row=1, column=10).fill = blue_fill
      worksheet.cell(row=1, column=11).fill = blue_fill
      worksheet.cell(row=1, column=12).fill = blue_fill
      worksheet.cell(row=1, column=13).fill = blue_fill
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(14)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(15)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(16)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(17)].hidden = True
      
      export_df_nationality_conflict_style.to_excel(writer, index=False, sheet_name='Nationality Conflict')
      worksheet = writer.sheets['Nationality Conflict']
      worksheet.cell(row=1, column=2).fill = orange_fill
      worksheet.cell(row=1, column=3).fill = orange_fill
      worksheet.cell(row=1, column=4).fill = orange_fill
      worksheet.cell(row=1, column=5).fill = orange_fill
      worksheet.cell(row=1, column=6).fill = orange_fill
      worksheet.cell(row=1, column=7).fill = orange_fill
      worksheet.cell(row=1, column=8).fill = blue_fill
      worksheet.cell(row=1, column=9).fill = blue_fill
      worksheet.cell(row=1, column=10).fill = blue_fill
      worksheet.cell(row=1, column=11).fill = blue_fill
      worksheet.cell(row=1, column=12).fill = blue_fill
      worksheet.cell(row=1, column=13).fill = blue_fill
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(14)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(15)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(16)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(17)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(18)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(19)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(20)].hidden = True
      worksheet.column_dimensions[openpyxl.utils.get_column_letter(21)].hidden = True
      
      export_df_parent_conflict_style.to_excel(writer, index=False, sheet_name='Parent Conflict')
      worksheet = writer.sheets['Parent Conflict']

      # Calculate the number of columns to fill based on the formula and fill them
      num_columns_to_fill = 2 + (max_i * 4) + 1
      for col_index in range(2, num_columns_to_fill + 1):
          worksheet.cell(row=1, column=col_index).fill = orange_fill
      
      num_columns_to_fill1 = 2 + (4 * 4) + 1
      for col_index in range(num_columns_to_fill + 1, num_columns_to_fill + num_columns_to_fill1):
          worksheet.cell(row=1, column=col_index).fill = blue_fill
      for col_index in range(num_columns_to_fill + num_columns_to_fill1,60):
          worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].hidden = True
  print('iSAMS-OA Synchronizing process is done.')
  
      
if __name__ == "__main__":
  isams_oa_sync()